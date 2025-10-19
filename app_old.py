from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, IntegerField, FloatField, TextAreaField, SelectField, SubmitField
from wtforms.validators import DataRequired, Email, Length, EqualTo, NumberRange
from datetime import datetime, timedelta
import os
import jwt
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io

# Load environment variables
load_dotenv()

# Flask App Configuration
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'celiker-enterprise-secret-key-2024')
app.config['WTF_CSRF_ENABLED'] = True

# Database Configuration
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Production PostgreSQL (Render.com)
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # Local development - SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///celiker_enterprise.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Initialize Extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmalısınız.'
login_manager.login_message_category = 'info'

# Enterprise Database Models

class User(UserMixin, db.Model):
    __tablename__ = 'users'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(128), nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    company = db.Column(db.String(100), nullable=True)
    role = db.Column(db.String(20), default='user')  # admin, manager, user
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    
    # Relationships
    products = db.relationship('Urun', backref='owner', lazy=True, cascade='all, delete-orphan')
    activities = db.relationship('UserActivity', backref='user', lazy=True, cascade='all, delete-orphan')
    
    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    
    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)
    
    def get_full_name(self):
        return f"{self.first_name} {self.last_name}"
    
    def generate_jwt_token(self):
        payload = {
            'user_id': self.id,
            'username': self.username,
            'exp': datetime.utcnow() + timedelta(hours=24)
        }
        return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')
    
    def __repr__(self):
        return f'<User {self.username}>'

class Urun(db.Model):
    __tablename__ = 'products'
    
    id = db.Column(db.Integer, primary_key=True)
    ad = db.Column(db.String(100), nullable=False, index=True)
    barkod = db.Column(db.String(50), nullable=False, index=True)
    stok_adedi = db.Column(db.Integer, nullable=False, default=0)
    birim_fiyat = db.Column(db.Float, nullable=False, default=0.0)
    kategori = db.Column(db.String(50), default='Genel', index=True)
    aciklama = db.Column(db.Text)
    min_stok_seviyesi = db.Column(db.Integer, default=10)
    max_stok_seviyesi = db.Column(db.Integer, default=1000)
    
    # User relationship
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    
    # Timestamps
    olusturma_tarihi = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    guncelleme_tarihi = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Unique constraint per user
    __table_args__ = (db.UniqueConstraint('barkod', 'user_id', name='unique_barcode_per_user'),)
    
    @property
    def toplam_deger(self):
        return self.stok_adedi * self.birim_fiyat
    
    @property
    def stok_durumu(self):
        if self.stok_adedi == 0:
            return 'kritik'
        elif self.stok_adedi <= self.min_stok_seviyesi:
            return 'dusuk'
        elif self.stok_adedi >= self.max_stok_seviyesi:
            return 'fazla'
        return 'normal'
    
    def to_dict(self):
        return {
            'id': self.id,
            'ad': self.ad,
            'barkod': self.barkod,
            'stok_adedi': self.stok_adedi,
            'birim_fiyat': self.birim_fiyat,
            'kategori': self.kategori,
            'aciklama': self.aciklama,
            'toplam_deger': self.toplam_deger,
            'stok_durumu': self.stok_durumu,
            'min_stok_seviyesi': self.min_stok_seviyesi,
            'max_stok_seviyesi': self.max_stok_seviyesi,
            'olusturma_tarihi': self.olusturma_tarihi.isoformat(),
            'guncelleme_tarihi': self.guncelleme_tarihi.isoformat()
        }
    
    def __repr__(self):
        return f'<Urun {self.ad}>'

class UserActivity(db.Model):
    __tablename__ = 'user_activities'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    action = db.Column(db.String(50), nullable=False)  # create, update, delete, login, logout
    resource_type = db.Column(db.String(50))  # product, user, etc.
    resource_id = db.Column(db.Integer)
    details = db.Column(db.JSON)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    
    def __repr__(self):
        return f'<Activity {self.action} by {self.user_id}>'

# Flask-Login user loader
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Enterprise Forms
class LoginForm(FlaskForm):
    username = StringField('Kullanıcı Adı', validators=[DataRequired(), Length(min=3, max=80)])
    password = PasswordField('Şifre', validators=[DataRequired()])
    submit = SubmitField('Giriş Yap')

class RegisterForm(FlaskForm):
    username = StringField('Kullanıcı Adı', validators=[DataRequired(), Length(min=3, max=80)])
    email = StringField('E-posta', validators=[DataRequired(), Email()])
    first_name = StringField('Ad', validators=[DataRequired(), Length(min=2, max=50)])
    last_name = StringField('Soyad', validators=[DataRequired(), Length(min=2, max=50)])
    company = StringField('Şirket', validators=[Length(max=100)])
    password = PasswordField('Şifre', validators=[DataRequired(), Length(min=6)])
    password2 = PasswordField('Şifre Tekrar', validators=[DataRequired(), EqualTo('password')])
    submit = SubmitField('Kayıt Ol')

class ProductForm(FlaskForm):
    ad = StringField('Ürün Adı', validators=[DataRequired(), Length(min=2, max=100)])
    barkod = StringField('Barkod', validators=[DataRequired(), Length(min=1, max=50)])
    stok_adedi = IntegerField('Stok Adedi', validators=[DataRequired(), NumberRange(min=0)])
    birim_fiyat = FloatField('Birim Fiyat (₺)', validators=[DataRequired(), NumberRange(min=0)])
    kategori = SelectField('Kategori', choices=[
        ('Genel', 'Genel'),
        ('Elektronik', 'Elektronik'),
        ('Gıda', 'Gıda'),
        ('Giyim', 'Giyim'),
        ('Ev & Yaşam', 'Ev & Yaşam'),
        ('Spor', 'Spor'),
        ('Kitap', 'Kitap'),
        ('Oyuncak', 'Oyuncak'),
        ('Diğer', 'Diğer')
    ], default='Genel')
    min_stok_seviyesi = IntegerField('Minimum Stok Seviyesi', validators=[NumberRange(min=0)], default=10)
    max_stok_seviyesi = IntegerField('Maksimum Stok Seviyesi', validators=[NumberRange(min=1)], default=1000)
    aciklama = TextAreaField('Açıklama', validators=[Length(max=500)])
    submit = SubmitField('Kaydet')

# Utility Functions
def log_user_activity(action, resource_type=None, resource_id=None, details=None):
    """Log user activity for audit trail"""
    if current_user.is_authenticated:
        activity = UserActivity(
            user_id=current_user.id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        db.session.add(activity)
        db.session.commit()

# Ana Sayfa
@app.route('/urun_ekle', methods=['GET', 'POST'])
def urun_ekle():
    if request.method == 'POST':
        ad = request.form['ad']
        barkod = request.form['barkod']
        stok_adedi = int(request.form['stok_adedi'])
        birim_fiyat = float(request.form['birim_fiyat'])
        kategori = request.form.get('kategori', 'Genel')
        aciklama = request.form.get('aciklama', '')
        
        # Barkod kontrolü
        mevcut_urun = Urun.query.filter_by(barkod=barkod).first()
        if mevcut_urun:
            flash('Bu barkod numarası zaten kullanılıyor!', 'error')
            return render_template('urun_ekle.html')
        
        yeni_urun = Urun(
            ad=ad,
            barkod=barkod,
            stok_adedi=stok_adedi,
            birim_fiyat=birim_fiyat,
            kategori=kategori,
            aciklama=aciklama
        )
        
        db.session.add(yeni_urun)
        db.session.commit()
        flash('Ürün başarıyla eklendi!', 'success')
        return redirect(url_for('index'))
    
    return render_template('urun_ekle.html')

# Ürün Düzenleme
@app.route('/urun_duzenle/<int:id>', methods=['GET', 'POST'])
def urun_duzenle(id):
    urun = Urun.query.get_or_404(id)
    
    if request.method == 'POST':
        urun.ad = request.form['ad']
        barkod = request.form['barkod']
        
        # Barkod kontrolü (kendi barkodu hariç)
        mevcut_urun = Urun.query.filter(Urun.barkod == barkod, Urun.id != id).first()
        if mevcut_urun:
            flash('Bu barkod numarası başka bir ürün tarafından kullanılıyor!', 'error')
            return render_template('urun_duzenle.html', urun=urun)
        
        urun.barkod = barkod
        urun.stok_adedi = int(request.form['stok_adedi'])
        urun.birim_fiyat = float(request.form['birim_fiyat'])
        urun.kategori = request.form.get('kategori', 'Genel')
        urun.aciklama = request.form.get('aciklama', '')
        
        db.session.commit()
        flash('Ürün başarıyla güncellendi!', 'success')
        return redirect(url_for('index'))
    
    return render_template('urun_duzenle.html', urun=urun)

# Ürün Silme
@app.route('/urun_sil/<int:id>')
def urun_sil(id):
    urun = Urun.query.get_or_404(id)
    db.session.delete(urun)
    db.session.commit()
    flash('Ürün başarıyla silindi!', 'success')
    return redirect(url_for('index'))

# Arama
@app.route('/ara')
def ara():
    arama_terimi = request.args.get('q', '')
    if arama_terimi:
        urunler = Urun.query.filter(
            (Urun.ad.contains(arama_terimi)) | 
            (Urun.barkod.contains(arama_terimi))
        ).all()
    else:
        urunler = []
    
    return render_template('arama_sonuclari.html', urunler=urunler, arama_terimi=arama_terimi)

# Barkod ile Ürün Arama (AJAX)
@app.route('/barkod_ara/<barkod>')
def barkod_ara(barkod):
    urun = Urun.query.filter_by(barkod=barkod).first()
    if urun:
        return jsonify(urun.to_dict())
    return jsonify({'error': 'Ürün bulunamadı'}), 404

# Excel Dışa Aktarma
@app.route('/excel_aktar')
def excel_aktar():
    urunler = Urun.query.all()
    
    # Workbook oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Listesi"
    
    # Başlık satırı
    basliklar = ['ID', 'Ürün Adı', 'Barkod', 'Stok Adedi', 'Birim Fiyat', 'Toplam Değer', 'Kategori', 'Açıklama', 'Oluşturma Tarihi']
    ws.append(basliklar)
    
    # Başlık stilini ayarla
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Veri satırları
    for urun in urunler:
        ws.append([
            urun.id,
            urun.ad,
            urun.barkod,
            urun.stok_adedi,
            urun.birim_fiyat,
            urun.toplam_deger,
            urun.kategori,
            urun.aciklama,
            urun.olusturma_tarihi.strftime('%d.%m.%Y %H:%M')
        ])
    
    # Sütun genişliklerini ayarla
    column_widths = [5, 25, 15, 12, 12, 15, 15, 30, 20]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    # Dosyayı kaydet
    filename = f"stok_listesi_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join('static', 'exports', filename)
    
    # Klasör yoksa oluştur
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# PDF Rapor Oluşturma
@app.route('/pdf_rapor')
def pdf_rapor():
    urunler = Urun.query.all()
    
    # PDF dosyası oluştur
    filename = f"stok_raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    filepath = os.path.join('static', 'exports', filename)
    
    # Klasör yoksa oluştur
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    story = []
    
    # Stil tanımlamaları
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Başlık
    title = Paragraph("STOK RAPORU", title_style)
    story.append(title)
    story.append(Spacer(1, 20))
    
    # Tarih
    date_text = f"Rapor Tarihi: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    date_para = Paragraph(date_text, styles['Normal'])
    story.append(date_para)
    story.append(Spacer(1, 20))
    
    # Tablo verisi
    data = [['Ürün Adı', 'Barkod', 'Stok', 'Birim Fiyat', 'Toplam Değer']]
    
    for urun in urunler:
        data.append([
            urun.ad,
            urun.barkod,
            str(urun.stok_adedi),
            f"{urun.birim_fiyat:.2f} ₺",
            f"{urun.toplam_deger:.2f} ₺"
        ])
    
    # Tablo oluştur
    table = Table(data, colWidths=[2*inch, 1.5*inch, 0.8*inch, 1*inch, 1.2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))
    
    # Özet bilgiler
    toplam_deger = sum(urun.toplam_deger for urun in urunler)
    ozet_text = f"Toplam Ürün Sayısı: {len(urunler)}<br/>Toplam Stok Değeri: {toplam_deger:.2f} ₺"
    ozet_para = Paragraph(ozet_text, styles['Normal'])
    story.append(ozet_para)
    
    doc.build(story)
    
    return send_file(filepath, as_attachment=True, download_name=filename)

# Düşük Stok Uyarısı
@app.route('/dusuk_stok')
def dusuk_stok():
    limit = request.args.get('limit', 10, type=int)
    urunler = Urun.query.filter(Urun.stok_adedi <= limit).all()
    return render_template('dusuk_stok.html', urunler=urunler, limit=limit)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    # Production için port ayarı
    import os
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
