from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, IntegerField, FloatField, TextAreaField, SelectField, SubmitField
from wtforms.validators import DataRequired, Email, Length, EqualTo, NumberRange
from datetime import datetime, timedelta
import os
import jwt
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
import io

# Load environment variables
load_dotenv()

# Flask App Configuration
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'celiker-enterprise-secret-key-2024')
app.config['WTF_CSRF_ENABLED'] = True

# Database Configuration
database_url = os.environ.get('DATABASE_URL')
if database_url:
    # Production PostgreSQL (Render.com)
    if database_url.startswith('postgres://'):
        database_url = database_url.replace('postgres://', 'postgresql://', 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    # Local development - SQLite
    app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///celiker_enterprise.db'

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {
    'pool_pre_ping': True,
    'pool_recycle': 300,
}

# Initialize Extensions
db = SQLAlchemy(app)
bcrypt = Bcrypt(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Bu sayfaya erişmek için giriş yapmalısınız.'
login_manager.login_message_category = 'info'

# Enterprise Database Models
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False, index=True)
    email = db.Column(db.String(120), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(128), nullable=False)
    first_name = db.Column(db.String(50), nullable=False)
    last_name = db.Column(db.String(50), nullable=False)
    company = db.Column(db.String(100), nullable=True)
    role = db.Column(db.String(20), default='user')  # admin, manager, user
    is_active = db.Column(db.Boolean, default=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    
    # Relationships
    products = db.relationship('Urun', backref='owner', lazy=True, cascade='all, delete-orphan')
    activities = db.relationship('UserActivity', backref='user', lazy=True, cascade='all, delete-orphan')
    
    def set_password(self, password):
        self.password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    
    def check_password(self, password):
        return bcrypt.check_password_hash(self.password_hash, password)
    
    def get_full_name(self):
        return f"{self.first_name} {self.last_name}"
    
    def generate_jwt_token(self):
        payload = {
            'user_id': self.id,
            'username': self.username,
            'exp': datetime.utcnow() + timedelta(hours=24)
        }
        return jwt.encode(payload, app.config['SECRET_KEY'], algorithm='HS256')
    
    def __repr__(self):
        return f'<User {self.username}>'

class Urun(db.Model):
    __tablename__ = 'products'
    
    id = db.Column(db.Integer, primary_key=True)
    ad = db.Column(db.String(100), nullable=False, index=True)
    barkod = db.Column(db.String(50), nullable=False, index=True)
    stok_adedi = db.Column(db.Integer, nullable=False, default=0)
    birim_fiyat = db.Column(db.Float, nullable=False, default=0.0)
    kategori = db.Column(db.String(50), default='Genel', index=True)
    aciklama = db.Column(db.Text)
    min_stok_seviyesi = db.Column(db.Integer, default=10)
    max_stok_seviyesi = db.Column(db.Integer, default=1000)
    
    # User relationship
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    
    # Timestamps
    olusturma_tarihi = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    guncelleme_tarihi = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    
    # Unique constraint per user
    __table_args__ = (db.UniqueConstraint('barkod', 'user_id', name='unique_barcode_per_user'),)
    
    @property
    def toplam_deger(self):
        return self.stok_adedi * self.birim_fiyat
    
    @property
    def stok_durumu(self):
        if self.stok_adedi == 0:
            return 'kritik'
        elif self.stok_adedi <= self.min_stok_seviyesi:
            return 'dusuk'
        elif self.stok_adedi >= self.max_stok_seviyesi:
            return 'fazla'
        return 'normal'
    
    def to_dict(self):
        return {
            'id': self.id,
            'ad': self.ad,
            'barkod': self.barkod,
            'stok_adedi': self.stok_adedi,
            'birim_fiyat': self.birim_fiyat,
            'kategori': self.kategori,
            'aciklama': self.aciklama,
            'toplam_deger': self.toplam_deger,
            'stok_durumu': self.stok_durumu,
            'min_stok_seviyesi': self.min_stok_seviyesi,
            'max_stok_seviyesi': self.max_stok_seviyesi,
            'olusturma_tarihi': self.olusturma_tarihi.isoformat(),
            'guncelleme_tarihi': self.guncelleme_tarihi.isoformat()
        }
    
    def __repr__(self):
        return f'<Urun {self.ad}>'

class UserActivity(db.Model):
    __tablename__ = 'user_activities'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False, index=True)
    action = db.Column(db.String(50), nullable=False)  # create, update, delete, login, logout
    resource_type = db.Column(db.String(50))  # product, user, etc.
    resource_id = db.Column(db.Integer)
    details = db.Column(db.JSON)
    ip_address = db.Column(db.String(45))
    user_agent = db.Column(db.Text)
    timestamp = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    
    def __repr__(self):
        return f'<Activity {self.action} by {self.user_id}>'

# Flask-Login user loader
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# Enterprise Forms
class LoginForm(FlaskForm):
    username = StringField('Kullanıcı Adı', validators=[DataRequired(), Length(min=3, max=80)])
    password = PasswordField('Şifre', validators=[DataRequired()])
    submit = SubmitField('Giriş Yap')

class RegisterForm(FlaskForm):
    username = StringField('Kullanıcı Adı', validators=[DataRequired(), Length(min=3, max=80)])
    email = StringField('E-posta', validators=[DataRequired(), Email()])
    first_name = StringField('Ad', validators=[DataRequired(), Length(min=2, max=50)])
    last_name = StringField('Soyad', validators=[DataRequired(), Length(min=2, max=50)])
    company = StringField('Şirket', validators=[Length(max=100)])
    password = PasswordField('Şifre', validators=[DataRequired(), Length(min=6)])
    password2 = PasswordField('Şifre Tekrar', validators=[DataRequired(), EqualTo('password')])
    submit = SubmitField('Kayıt Ol')

class ProductForm(FlaskForm):
    ad = StringField('Ürün Adı', validators=[DataRequired(), Length(min=2, max=100)])
    barkod = StringField('Barkod', validators=[DataRequired(), Length(min=1, max=50)])
    stok_adedi = IntegerField('Stok Adedi', validators=[DataRequired(), NumberRange(min=0)])
    birim_fiyat = FloatField('Birim Fiyat (₺)', validators=[DataRequired(), NumberRange(min=0)])
    kategori = SelectField('Kategori', choices=[
        ('Genel', 'Genel'),
        ('Elektronik', 'Elektronik'),
        ('Gıda', 'Gıda'),
        ('Giyim', 'Giyim'),
        ('Ev & Yaşam', 'Ev & Yaşam'),
        ('Spor', 'Spor'),
        ('Kitap', 'Kitap'),
        ('Oyuncak', 'Oyuncak'),
        ('Diğer', 'Diğer')
    ], default='Genel')
    min_stok_seviyesi = IntegerField('Minimum Stok Seviyesi', validators=[NumberRange(min=0)], default=10)
    max_stok_seviyesi = IntegerField('Maksimum Stok Seviyesi', validators=[NumberRange(min=1)], default=1000)
    aciklama = TextAreaField('Açıklama', validators=[Length(max=500)])
    submit = SubmitField('Kaydet')

# Utility Functions
def log_user_activity(action, resource_type=None, resource_id=None, details=None):
    """Log user activity for audit trail"""
    if current_user.is_authenticated:
        activity = UserActivity(
            user_id=current_user.id,
            action=action,
            resource_type=resource_type,
            resource_id=resource_id,
            details=details,
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        db.session.add(activity)
        db.session.commit()

# Authentication Routes
@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        
        if user and user.check_password(form.password.data) and user.is_active:
            login_user(user, remember=True)
            user.last_login = datetime.utcnow()
            db.session.commit()
            
            log_user_activity('login')
            flash(f'Hoş geldiniz, {user.get_full_name()}!', 'success')
            
            next_page = request.args.get('next')
            return redirect(next_page) if next_page else redirect(url_for('dashboard'))
        else:
            flash('Geçersiz kullanıcı adı veya şifre!', 'danger')
    
    return render_template('auth/login.html', form=form)

@app.route('/register', methods=['GET', 'POST'])
def register():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    form = RegisterForm()
    if form.validate_on_submit():
        # Check if username or email already exists
        existing_user = User.query.filter(
            (User.username == form.username.data) | 
            (User.email == form.email.data)
        ).first()
        
        if existing_user:
            if existing_user.username == form.username.data:
                flash('Bu kullanıcı adı zaten kullanılıyor!', 'danger')
            else:
                flash('Bu e-posta adresi zaten kayıtlı!', 'danger')
        else:
            user = User(
                username=form.username.data,
                email=form.email.data,
                first_name=form.first_name.data,
                last_name=form.last_name.data,
                company=form.company.data
            )
            user.set_password(form.password.data)
            
            db.session.add(user)
            db.session.commit()
            
            log_user_activity('register', 'user', user.id)
            flash('Kayıt başarılı! Şimdi giriş yapabilirsiniz.', 'success')
            return redirect(url_for('login'))
    
    return render_template('auth/register.html', form=form)

@app.route('/logout')
@login_required
def logout():
    log_user_activity('logout')
    logout_user()
    flash('Başarıyla çıkış yaptınız.', 'info')
    return redirect(url_for('login'))

# Dashboard (Ana Sayfa - Login Required)
@app.route('/dashboard')
@app.route('/')
@login_required
def dashboard():
    # User-specific data
    user_products = Urun.query.filter_by(user_id=current_user.id).all()
    
    # Statistics
    toplam_urun_sayisi = len(user_products)
    toplam_stok_degeri = sum(urun.toplam_deger for urun in user_products)
    dusuk_stoklu_urunler = len([p for p in user_products if p.stok_durumu in ['kritik', 'dusuk']])
    kritik_stoklu_urunler = len([p for p in user_products if p.stok_durumu == 'kritik'])
    
    # Category statistics
    kategori_stats = {}
    for urun in user_products:
        if urun.kategori not in kategori_stats:
            kategori_stats[urun.kategori] = {'count': 0, 'value': 0}
        kategori_stats[urun.kategori]['count'] += 1
        kategori_stats[urun.kategori]['value'] += urun.toplam_deger
    
    istatistikler = {
        'toplam_urun_sayisi': toplam_urun_sayisi,
        'toplam_stok_degeri': toplam_stok_degeri,
        'dusuk_stoklu_urunler': dusuk_stoklu_urunler,
        'kritik_stoklu_urunler': kritik_stoklu_urunler,
        'kategori_stats': kategori_stats
    }
    
    # Recent products
    son_urunler = Urun.query.filter_by(user_id=current_user.id)\
        .order_by(Urun.olusturma_tarihi.desc()).limit(5).all()
    
    # Low stock alerts
    dusuk_stok_urunler = [p for p in user_products if p.stok_durumu in ['kritik', 'dusuk']][:5]
    
    return render_template('dashboard.html', 
                         istatistikler=istatistikler, 
                         son_urunler=son_urunler,
                         dusuk_stok_urunler=dusuk_stok_urunler)

# Product Management Routes
@app.route('/urun_ekle', methods=['GET', 'POST'])
@login_required
def urun_ekle():
    form = ProductForm()
    if form.validate_on_submit():
        # Check if barcode already exists for this user
        existing_product = Urun.query.filter_by(
            barkod=form.barkod.data, 
            user_id=current_user.id
        ).first()
        
        if existing_product:
            flash('Bu barkod numarası zaten kullanılıyor!', 'danger')
        else:
            product = Urun(
                ad=form.ad.data,
                barkod=form.barkod.data,
                stok_adedi=form.stok_adedi.data,
                birim_fiyat=form.birim_fiyat.data,
                kategori=form.kategori.data,
                min_stok_seviyesi=form.min_stok_seviyesi.data,
                max_stok_seviyesi=form.max_stok_seviyesi.data,
                aciklama=form.aciklama.data,
                user_id=current_user.id
            )
            
            db.session.add(product)
            db.session.commit()
            
            log_user_activity('create', 'product', product.id, {
                'product_name': product.ad,
                'barcode': product.barkod
            })
            
            flash(f'Ürün "{product.ad}" başarıyla eklendi!', 'success')
            return redirect(url_for('dashboard'))
    
    return render_template('urun_ekle.html', form=form)

@app.route('/urun_listesi')
@login_required
def urun_listesi():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # User-specific products with pagination
    products = Urun.query.filter_by(user_id=current_user.id)\
        .order_by(Urun.guncelleme_tarihi.desc())\
        .paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template('urun_listesi.html', products=products)

@app.route('/urun_duzenle/<int:id>', methods=['GET', 'POST'])
@login_required
def urun_duzenle(id):
    product = Urun.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    
    form = ProductForm(obj=product)
    if form.validate_on_submit():
        # Check if barcode already exists for this user (excluding current product)
        existing_product = Urun.query.filter(
            Urun.barkod == form.barkod.data,
            Urun.user_id == current_user.id,
            Urun.id != id
        ).first()
        
        if existing_product:
            flash('Bu barkod numarası başka bir ürün tarafından kullanılıyor!', 'danger')
        else:
            old_data = product.to_dict()
            
            form.populate_obj(product)
            product.guncelleme_tarihi = datetime.utcnow()
            
            db.session.commit()
            
            log_user_activity('update', 'product', product.id, {
                'old_data': old_data,
                'new_data': product.to_dict()
            })
            
            flash(f'Ürün "{product.ad}" başarıyla güncellendi!', 'success')
            return redirect(url_for('urun_listesi'))
    
    return render_template('urun_duzenle.html', form=form, product=product)

@app.route('/urun_sil/<int:id>', methods=['POST'])
@login_required
def urun_sil(id):
    product = Urun.query.filter_by(id=id, user_id=current_user.id).first_or_404()
    
    log_user_activity('delete', 'product', product.id, {
        'product_name': product.ad,
        'barcode': product.barkod
    })
    
    db.session.delete(product)
    db.session.commit()
    
    flash(f'Ürün "{product.ad}" başarıyla silindi!', 'success')
    return redirect(url_for('urun_listesi'))

# Search and Filter Routes
@app.route('/ara')
@login_required
def ara():
    query = request.args.get('q', '')
    kategori = request.args.get('kategori', '')
    stok_durumu = request.args.get('stok_durumu', '')
    
    # Base query for user's products
    products_query = Urun.query.filter_by(user_id=current_user.id)
    
    # Apply filters
    if query:
        products_query = products_query.filter(
            (Urun.ad.contains(query)) | 
            (Urun.barkod.contains(query)) |
            (Urun.aciklama.contains(query))
        )
    
    if kategori:
        products_query = products_query.filter_by(kategori=kategori)
    
    products = products_query.order_by(Urun.guncelleme_tarihi.desc()).all()
    
    # Filter by stock status (done in Python for complex logic)
    if stok_durumu:
        products = [p for p in products if p.stok_durumu == stok_durumu]
    
    return render_template('arama_sonuclari.html', 
                         products=products, 
                         query=query,
                         kategori=kategori,
                         stok_durumu=stok_durumu)

@app.route('/dusuk_stok')
@login_required
def dusuk_stok():
    # Get products with low or critical stock
    products = Urun.query.filter_by(user_id=current_user.id).all()
    dusuk_stok_products = [p for p in products if p.stok_durumu in ['kritik', 'dusuk']]
    
    return render_template('dusuk_stok.html', products=dusuk_stok_products)

# Missing Routes for compatibility
@app.route('/pdf_rapor')
@login_required
def pdf_rapor():
    # Placeholder for PDF report - will implement later
    flash('PDF rapor özelliği yakında eklenecek!', 'info')
    return redirect(url_for('dashboard'))

# Export Routes
@app.route('/excel_aktar')
@login_required
def excel_aktar():
    # User-specific products
    products = Urun.query.filter_by(user_id=current_user.id).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Stok Listesi"
    
    # Headers
    headers = ['ID', 'Ürün Adı', 'Barkod', 'Stok Adedi', 'Birim Fiyat (₺)', 
               'Toplam Değer (₺)', 'Kategori', 'Stok Durumu', 'Açıklama']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Data rows
    for product in products:
        ws.append([
            product.id,
            product.ad,
            product.barkod,
            product.stok_adedi,
            product.birim_fiyat,
            product.toplam_deger,
            product.kategori,
            product.stok_durumu,
            product.aciklama or ''
        ])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to memory
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    log_user_activity('export', 'excel', None, {'product_count': len(products)})
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'stok_listesi_{current_user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

# Initialize database - Flask 2.3+ compatible
def create_tables():
    with app.app_context():
        db.create_all()

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    # Production için port ayarı
    port = int(os.environ.get('PORT', 5000))
    debug_mode = os.environ.get('FLASK_ENV') != 'production'
    
    app.run(debug=debug_mode, host='0.0.0.0', port=port)
